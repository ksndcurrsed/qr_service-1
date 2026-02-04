import time
import asyncio
import aiohttp
import ssl
import os
import datetime
import win32print
import win32ui
import win32con
from PIL import Image, ImageWin
from pylibdmtx.pylibdmtx import encode
from pynput import keyboard
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from dotenv import load_dotenv

load_dotenv()

# --- НАСТРОЙКИ ---
SERVER_IP = "84.54.29.24"
SERVER_DOMAIN = "fffzar-tool.ru"
# Используем HTTPS после настройки SSL
SERVER_URL = f"https://{SERVER_DOMAIN}"
HISTORY_FOLDER = "history"
# При каждом запуске — новый отчёт: report_дд-мм-гггг_чч-мм.xlsx
EXCEL_FILE = f"report_{datetime.datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"

# Создаем папку для истории, если её нет
if not os.path.exists(HISTORY_FOLDER):
    os.makedirs(HISTORY_FOLDER)

def save_to_report(text, file_path):
    """Записывает данные о сканировании в Excel"""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата и время", "Содержимое кода", "Имя файла"])
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

    ws.append([timestamp, ILLEGAL_CHARACTERS_RE.sub(r'', text), os.path.basename(file_path)])
    wb.save(EXCEL_FILE)

# Размер этикетки 58×40 мм, DPI термопринтера (обычно 203)
LABEL_MM = (58, 40)
PRINTER_DPI = 203
QR_FILL_RATIO = float(os.environ.get("QR_FILL_RATIO", "0.82"))  # было 0.9; 0.82 безопаснее для полей/перекоса ленты


def mm_to_px(mm, dpi=PRINTER_DPI):
    """Конвертация миллиметров в пиксели принтера"""
    return int(mm / 25.4 * dpi)


# --- Печать: пытаемся зафиксировать формат 58x40 в драйвере и всегда опираемся на реальные размеры DC ---
def _dmpaper_user_const() -> int:
    # В некоторых сборках pywin32 константа может отсутствовать
    return int(getattr(win32con, "DMPAPER_USER", 256))


def build_label_devmode(printer_name: str, width_mm: int, height_mm: int):
    """
    Возвращает DEVMODE с кастомным размером бумаги (в 0.1 мм), если драйвер это поддерживает.
    Даже если драйвер проигнорирует, дальше мы всё равно используем реальные DeviceCaps.
    """
    try:
        hPrinter = win32print.OpenPrinter(printer_name)
        try:
            info2 = win32print.GetPrinter(hPrinter, 2)
            devmode = info2.get("pDevMode")
            if devmode is None:
                return None

            devmode.Fields |= (
                win32con.DM_PAPERSIZE
                | win32con.DM_PAPERWIDTH
                | win32con.DM_PAPERLENGTH
                | win32con.DM_ORIENTATION
            )
            devmode.PaperSize = _dmpaper_user_const()
            devmode.PaperWidth = int(width_mm * 10)   # 0.1 мм
            devmode.PaperLength = int(height_mm * 10) # 0.1 мм
            devmode.Orientation = (
                win32con.DMORIENT_LANDSCAPE
                if width_mm >= height_mm
                else win32con.DMORIENT_PORTRAIT
            )
            return devmode
        finally:
            win32print.ClosePrinter(hPrinter)
    except Exception as e:
        print(f"Не удалось подготовить настройки бумаги принтера (DEVMODE): {e}")
        return None


def create_printer_dc(printer_name: str, label_mm: tuple[int, int] = LABEL_MM):
    """
    Создаёт DC принтера; пытается выставить нужный размер этикетки через DEVMODE.
    """
    devmode = build_label_devmode(printer_name, label_mm[0], label_mm[1])
    hDC = win32ui.CreateDC()

    if devmode is not None:
        try:
            # CreateDC(driver, device, output, devmode)
            hDC.CreateDC("WINSPOOL", printer_name, None, devmode)
            return hDC
        except Exception as e:
            print(f"DEVMODE не применился, печатаю с настройками драйвера по умолчанию: {e}")

    hDC.CreatePrinterDC(printer_name)
    return hDC


def get_dc_page_px(hDC) -> tuple[int, int]:
    """
    Размер печатаемой области (в пикселях устройства) для текущих настроек драйвера.
    """
    w = int(hDC.GetDeviceCaps(win32con.HORZRES) or 0)
    h = int(hDC.GetDeviceCaps(win32con.VERTRES) or 0)
    return w, h


# Защита от дублей: один и тот же код не печатаем повторно в течение 5 сек
_last_printed = {}
_DEDUP_SEC = 2


def process_and_print(text):
    """Генерация, сохранение, логирование и печать"""
    global _last_printed
    now = time.time()
    if text in _last_printed and (now - _last_printed[text]) < _DEDUP_SEC:
        print(f"Пропуск дубля: {text[:30]}...")
        return
    _last_printed[text] = now
    # Очистка старых записей
    _last_printed = {k: v for k, v in _last_printed.items() if now - v < 60}

    print(f"Обработка: {text}")
    try:
        # 1. Генерация Data Matrix
        encoded = encode(text.encode('utf-8'))
        img = Image.frombytes('RGB', (encoded.width, encoded.height), encoded.pixels)

        # 2. Создаём DC заранее и берём РЕАЛЬНЫЙ размер страницы у принтера
        printer_name = win32print.GetDefaultPrinter()
        hDC = create_printer_dc(printer_name, LABEL_MM)
        try:
            page_w, page_h = get_dc_page_px(hDC)

            # Фолбэк, если драйвер вернул 0 (редко, но бывает)
            if page_w <= 0 or page_h <= 0:
                dpi_x = int(hDC.GetDeviceCaps(win32con.LOGPIXELSX) or PRINTER_DPI)
                dpi_y = int(hDC.GetDeviceCaps(win32con.LOGPIXELSY) or PRINTER_DPI)
                page_w = mm_to_px(LABEL_MM[0], dpi_x)
                page_h = mm_to_px(LABEL_MM[1], dpi_y)

            # Лог полезен для диагностики "вдруг драйвер сменил бумагу/ориентацию"
            try:
                dpi_x = int(hDC.GetDeviceCaps(win32con.LOGPIXELSX) or 0)
                dpi_y = int(hDC.GetDeviceCaps(win32con.LOGPIXELSY) or 0)
                print(f"Принтер: {printer_name} | page={page_w}x{page_h}px | dpi={dpi_x}x{dpi_y}")
            except Exception:
                pass

            # 3. QR/DM занимает заданную долю меньшей стороны, по центру
            qr_size = max(1, int(min(page_w, page_h) * QR_FILL_RATIO))
            img = img.resize((qr_size, qr_size), Image.NEAREST)

            # 4. Создаём полную этикетку с центрированным QR
            label_img = Image.new('RGB', (page_w, page_h), 'white')
            x = max(0, (page_w - qr_size) // 2)
            y = max(0, (page_h - qr_size) // 2)
            label_img.paste(img, (x, y))

            # 5. Сохранение в папку истории с уникальным именем
            file_name = f"scan_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.png"
            full_path = os.path.join(HISTORY_FOLDER, file_name)
            label_img.save(full_path)

            # 6. Запись в Excel
            save_to_report(text, full_path)

            # 7. Печать всей этикетки (ровно в размеры страницы DC)
            hDC.StartDoc("DM_Job")
            hDC.StartPage()
            dib = ImageWin.Dib(label_img)
            dib.draw(hDC.GetHandleOutput(), (0, 0, page_w, page_h))
            hDC.EndPage()
            hDC.EndDoc()
            print(f"Успешно: данные сохранены и отправлены на принтер.")
        finally:
            try:
                hDC.DeleteDC()
            except Exception:
                pass
        
    except Exception as e:
        print(f"Ошибка при обработке: {e}")

# --- Логика перехвата (USB/Bluetooth сканер) ---
buffer = []

last_key_time = 0
is_scanner_typing = True

def on_press(key):
    global buffer, last_key_time, is_scanner_typing
    
    current_time = time.time()
    # Считаем время с момента прошлого нажатия
    delay = current_time - last_key_time
    last_key_time = current_time

    # Если пауза между буквами больше 50мс — это скорее всего человек
    if delay > 0.05:
        is_scanner_typing = False

    try:
        if key == keyboard.Key.enter:
            msg = "".join(buffer)
            if is_scanner_typing and len(msg) > 15:
                process_and_print(msg)

            buffer = []
            is_scanner_typing = True 
        elif hasattr(key, 'char') and key.char:
            buffer.append(key.char)
    except:
        pass

# --- Связь с сервером (HTTP API) ---
async def listen():
    """Опрос сервера на наличие новых заданий для печати"""
    # Настройка SSL для работы в exe (отключаем проверку сертификата)
    ssl_context = ssl.create_default_context()
    ssl_context.check_hostname = False
    ssl_context.verify_mode = ssl.CERT_NONE
    
    # Создаем connector с SSL контекстом
    connector = aiohttp.TCPConnector(ssl=ssl_context)
    
    async with aiohttp.ClientSession(connector=connector) as session:
        print(f"Подключение к серверу: {SERVER_URL}")
        while True:
            try:
                async with session.get(f"{SERVER_URL}/get-job", timeout=aiohttp.ClientTimeout(total=10)) as response:
                    if response.status == 200:
                        data = await response.json()
                        if data.get("status") == "ok" and data.get("data"):
                            process_and_print(data["data"])
                            print(f"Получено задание: {data['data']}")
                    await asyncio.sleep(2)  # Проверка каждые 2 секунды
            except asyncio.TimeoutError:
                print("Таймаут подключения. Повтор через 5 сек...")
                await asyncio.sleep(5)
            except Exception as e:
                print(f"Ошибка подключения к серверу: {e}. Повтор через 5 сек...")
                await asyncio.sleep(5)

if __name__ == "__main__":
    keyboard.Listener(on_press=on_press).start()
    asyncio.run(listen())